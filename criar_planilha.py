from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from main import FILENAME

wb= Workbook()

def criar_planilha():
    planilha = wb.worksheets[0]

    planilha['A1'] = "Ethereum"; planilha['A2'] = "Revendedora"; planilha['B2'] = "placa de video"; planilha['C2'] = "Investimento Placa (reais)"
    planilha['D2'] = "Preço energia (kWh em dolar)"; planilha['E2'] = "Custo energia/mes"; planilha['F2']= "Rendimento 100(%) eficiencia"; planilha ['G2'] = "Lucro 100(%) eficiencia/placa/mes"
    planilha['H2'] = "Valor dolar"; planilha['I2'] = "depreciacao placa de video/ano "; planilha['J2'] = "depreciacao placa de video/mes"; planilha ['K2'] = "tempo retorno do investimento meses"
    planilha['L2'] = "quantidade placas"; planilha['M2'] = "Lucro/ano"; planilha['N2'] = "Investimento inicial por placa"; planilha['O2'] = "ROIC por placa"; planilha['P2'] = "investimentos extras"
    planilha['P4'] = "armacao do rig"; planilha['P6'] = "Placa mae"; planilha['P8'] = "fonte de alimentacao(PCU)"; planilha['P10'] = "CPU"; planilha['P12'] = "Memoria RAM(4 ou 8GB)"; planilha['P14'] = "SSD"
    planilha['P16'] = "PCI-e Riser"; planilha["Q2"] = "investimento total"; planilha['R2'] = "Lucro geral/ ano"; planilha['S2'] = "ROIC (%)"; planilha['I27'] = "fonte:tech spot"; planilha['G27'] = "fonte: what to mine"
    planilha['F27'] = "fonte: what to mine"; planilha['E27'] = "fonte: what to mine"; planilha["A28"] = "Valor energia / kWh reais"; planilha['B3'] = "RTX 2060"; planilha['B4'] = "RTX 2070"; planilha['B5'] = "RTX 2080"
    planilha['B6'] = "RTX 2080 Ti"; planilha['B7'] = "RTX 3080"; planilha['B8'] = "RTX 3090"; planilha['B9'] = "RTX 3060 Ti"; planilha['B10'] = "RTX 3070"; planilha['B11'] = "GTX 1660S"; planilha['B12'] = "GTX 1660 Ti"
    planilha['B13'] = "GTX 1660"; planilha['B14'] = "GTX 1080 Ti"; planilha['B15'] = "GTX 1080"; planilha['B16'] = "6800 XT"; planilha['B17'] = "6800"; planilha['B18'] = "VII"; planilha['B19'] = "5700 XT"
    planilha['B20'] = "5700"; planilha['B21'] = "5600 XT"; planilha['B22'] = "Vega64"; planilha['B23'] = "Vega56"; planilha['B24'] = "580"; planilha['B25'] = "570"; planilha['B26'] = "480"
    
    planilha['F3'] = "=F29*H3"
    for x in range(3,27):
        planilha.cell(row=x, column=6, value=Translator("=F29*H3", origin="F3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['G3'] = "=G29*H3"
    for x in range(3,27):
        planilha.cell(row=x, column=7, value=Translator("=G29*H3", origin="G3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['E3'] = "=F3-G3"
    for x in range(3,27):
        planilha.cell(row=x, column=5, value=Translator("=F3-G3", origin="E3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['D3'] = "=A29/H3"
    for x in range(3,27):
        planilha.cell(row=x, column=4, value=Translator("=A29/H3", origin="D3").translate_formula(row_delta=x-3,col_delta=0))
    for x in range(3,27):
        planilha.cell(row=x, column=9, value=0.15)
    planilha['J3'] = "=I3/12"
    for x in range(3,27):
        planilha.cell(row=x, column=10, value=Translator("=I3/12", origin="J3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['K3'] = "=C3/(G3-(J3*C3))"
    for x in range(3,27):
        planilha.cell(row=x, column=11, value=Translator("=C3/(G3-(J3*C3))", origin="K3").translate_formula(row_delta=x-3,col_delta=0))
    for x in range(3,27):
        planilha.cell(row=x, column=12, value=1)
    planilha['M3'] = "=((C3*12)/K3)*L3"
    for x in range(3,27):
        planilha.cell(row=x, column=13, value=Translator("=((C3*12)/K3)*L3", origin="M3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['N3'] = "=C3*L3"
    for x in range(3,27):
        planilha.cell(row=x, column=14, value=Translator("=C3*L3", origin="N3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['O3'] = "=(12/K3)*100"
    for x in range(3,27):
        planilha.cell(row=x, column=15, value=Translator("=(12/K3)*100", origin="O3").translate_formula(row_delta=x-3,col_delta=0))
    planilha['P3'] = "=SOMA(P4:P26)"; planilha['P5'] = 300; planilha['P7'] = 550; planilha['P9'] = 600; planilha['P11'] = 150
    planilha['P13'] = 100; planilha['P15'] = 100
    planilha['Q3'] = "=P3+SOMA(N3:N26)"
    planilha['R3'] = "=SOMA(M3:M26)"
    planilha['S3'] = "=(R3/Q3)*100"
    print("\nplanilha criada com sucesso")        
    wb.save(FILENAME)


criar_planilha()
    
