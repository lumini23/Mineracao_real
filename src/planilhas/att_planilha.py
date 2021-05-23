from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import math
from variables import FILENAME, DATA,URLS

def arredondar(num):
    return float( '%g' % (num))

def number_array_to_value(array,posicao):
    letra = 0
    first = False
    for x in range(posicao):
        array[x] = array[x]*pow(10,posicao - 1 - x)
    for x in range(posicao + 1,len(array)):
        if (type(array[x]) == float) or (type(array[x]) == int):
            array[x] = array[x]*pow(10,posicao - x)
        elif (type(array[x]) != float) and (type(array[x]) != int) and first == False:
            letra = x
            first = True
            x = len(array)
    if letra == 0:
        valor = 0.0
        for z in range(len(array)):
            if z != posicao:
                valor = arredondar(valor + array[z])
    else:
        valor = 0.0
        for z in range(letra):
            if z != posicao:
                valor = arredondar(valor + array[z])  
    return valor

def array_to_value(array):
    x = 0
    ponto = False 
    count = 0
    while x < len(array) and count < 3:
        if array[x] == '.':
            ponto = True
            y = x
        if ponto == True:
            count = count + 1
        if count != 1:
            array[x] = float(array[x])
        x = x + 1
    valor2060 = number_array_to_value(array,y)
    return valor2060

def getDolar():
        URL = "https://www.google.com/search?q=valor+dolar&oq=valor+dolar&aqs=chrome..69i57j35i39j0l8.6964j1j9&sourceid=chrome&ie=UTF-8"
        headers = {'User-Agent': "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.192 Safari/537.36"}
        site = requests.get(URL, headers=headers)
        soup = BeautifulSoup(site.content,'html.parser')
        valor = soup.find('span', class_="DFlfde SwHCTb").get_text().strip()
        num_valor = valor[0:4]
        num_valor = num_valor.replace(',','.')
        num_valor = float(num_valor)
        return num_valor

class whattomine:
    def __init__(self):
        self.urls = URLS
    def get_tabela_whattomine(self,URL):
        headers = {'User-Agent': "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.192 Safari/537.36"}
        site = requests.get(URL, headers=headers)
        soup = BeautifulSoup(site.content,'html.parser')
        tabela = soup.find('tr', class_="table-success").get_text().strip()
        return tabela
    def whattomineRent(self,x):
        URL = self.urls[x]['url']
        tabela = wtm.get_tabela_whattomine(URL)
        rentabilidade = ["A"]*7
        rent = True
        for x in range(190,220):
            if ((tabela[x] == "$") and (rent == True)):
                for y in range(5):
                    rentabilidade[y] = tabela[x+y+1]
                rent = False
        valor = array_to_value(rentabilidade)*30
        return valor
    def whattomineLucro(self,x):
        URL = self.urls[x]['url']
        tabela = wtm.get_tabela_whattomine(URL)
        Lucro = ["A"]*7
        count = 0
        for x in range(190,230):
            if tabela[x] == "$":
                count = count + 1
            if count == 2:
                for y in range(5):
                    Lucro[y] = tabela[x+y+1]
                count = 10
        valor = array_to_value(Lucro)*30
        return valor 
wtm = whattomine()

class planilha(whattomine):
    def __init__(self):
        self.wb = load_workbook(FILENAME)
    def atualizarRentabilidade(self):
        planilha = self.wb.worksheets[0]
        for x in range(29,53):
            planilha.cell(row=x,column=6,value=wtm.whattomineRent(x-29))
        print(" Rentabilidade da planilha atualizada com sucesso")
        self.wb.save(FILENAME)
    
    def atualizarDolar(self):
        planilha = self.wb.worksheets[0]
        dolar = getDolar()
        for x in range(3,27):
            planilha.cell(row=x, column=8, value=dolar)
        print(" Valor dolar atualizado com sucesso")
        self.wb.save(FILENAME)
    
    def atualizarLucro(self):
        planilha = self.wb.worksheets[0]
        for x in range(29,53):
            planilha.cell(row=x,column=7,value=wtm.whattomineLucro(x-29))
        print(" Valor do lucro atualizado com sucesso")
        self.wb.save(FILENAME)

pl = planilha()
pl.atualizarDolar()
pl.atualizarLucro()
pl.atualizarRentabilidade()
