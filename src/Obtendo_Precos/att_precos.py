from Pichau import PICHAU
from planilhas.variables import URLS,FILENAME
from openpyxl import load_workbook
import smtplib
import email.message
import xlwings as xw
from Magalu import MAGALU

wbxl=xw.Book(FILENAME)
wb = load_workbook(FILENAME)
planilha = wb.worksheets[0]
values = []
array = []
for row in planilha.iter_rows(min_row=3,min_col=3, max_col=3, max_row=26, values_only=True):
    values.append(row)
for x in range(len(values)):
    array.append(float(values[x][0]))

def atualizando_precos(DATA):
    for item in DATA:
        if DATA[item]['preco'] != "Produto indisponivel":
            for x in URLS:
                if DATA[item]['nome'] == URLS[x]['placa']:
                    if DATA[item]['preco'] < array[x]:
                        planilha.cell(row=x+3,column=3,value=DATA[item]['preco'])
    print("atualizacao concluida")
    wb.save(FILENAME)

def send_email():
    email_content = "URL"
    msg = email.message.Message()
    msg['Subject'] = 'Preco baixou'
    msg['From'] = 'danielmattarteste@gmail.com'
    msg['To'] = 'danielmattar2001@gmail.com'
    password = 'ggsx tutb jvhz iaeo'
    msg.add_header('Content-Type', 'text/html')
    msg.set_payload(email_content)

    s = smtplib.SMTP('smtp.gmail.com: 587')
    s.starttls()
    s.login(msg['From'],password)
    s.sendmail(msg['From'], [msg['To']], msg.as_string())

def checkROIC():
    ROIC = wbxl.sheets[0].range('O3','O26').value
    for x in range(len(ROIC)):
        if ROIC[x] > 80.0:
            print("aqui")
            send_email()
atualizando_precos(MAGALU)
send_email()
